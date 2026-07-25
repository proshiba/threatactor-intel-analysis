#!/usr/bin/env python3
"""Build an evidence-backed census of every threat actor named in the corpus."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from bootstrap_all_profiles import (
    actor_name_cells,
    load_workbook_rows,
    normalized_name,
)
from common import load_json, slugify, stable_digest, utc_now, write_json_atomic
from ingest_observables import SUPPORTED_SUFFIXES, source_records


IDENTIFIER_RE = re.compile(
    r"(?<![A-Z0-9-])(?:APT[- ]?\d{1,3}|UNC\d{3,4}|UNG\d{3,4}|UTG\d{3,4}|"
    r"FIN\d{1,4}|TA\d{3,4}|UAC-\d{4}|DEV-\d{4}|TAG-\d{2,4}|"
    r"Storm-\d{4}|CL-STA-\d{4})(?![A-Z0-9-])",
    re.IGNORECASE,
)
GENERIC_NAMES = {
    "actor", "actors", "adversary", "campaign", "china", "group", "iran",
    "malware", "north korea", "operation", "russia", "team", "threat actor",
    "unknown", "uncategorized",
}
SKIP_DIRS = {
    ".git", "profiles", "actor_profile", "data", "__pycache__", ".agents", ".codex",
}
AMBIGUOUS_SINGLE_NAMES = {
    "lead", "play", "silence", "equation", "bitter", "careto", "tick",
    "platinum", "patchwork", "anonymous", "carbanak",
}


def clean_name(value: str) -> str:
    return " ".join(value.replace("\u0000", " ").split()).strip(" ,;/")


def useful_name(value: str) -> bool:
    value = clean_name(value)
    return (
        len(value) >= 4
        and normalized_name(value) not in {normalized_name(x) for x in GENERIC_NAMES}
        and not value.lower().startswith(("http://", "https://"))
    )


def split_alias_cell(value: str) -> list[str]:
    return [
        clean_name(item)
        for item in re.split(r"[,;\n]|\s+/\s+", value)
        if useful_name(item)
    ]


def workbook_actor_names(record: dict[str, Any]) -> list[str]:
    """Return naming-taxonomy cells, excluding origin/target/overlap prose."""
    excluded = (
        "operation", "toolset", "malware", "target", "modus", "comment",
        "link", "mitre", "origin", "country", "region", "sponsor",
        "attribution", "motivation", "overlap", "associated",
    )
    names: list[str] = []
    for header, value in record["fields"].items():
        lower = header.casefold()
        if any(marker in lower for marker in excluded):
            continue
        if lower == "other names":
            names.extend(split_alias_cell(value))
        elif useful_name(value):
            names.append(clean_name(value))
    return list(dict.fromkeys(names))


def identity_id(mitre_id: str | None, name: str) -> str:
    return (
        f"actor-census--mitre:{mitre_id}"
        if mitre_id
        else f"actor-census--sha256:{stable_digest(normalized_name(name))}"
    )


def add_identity(
    identities: dict[str, dict[str, Any]],
    alias_index: dict[str, set[str]],
    *,
    name: str,
    aliases: Iterable[str] = (),
    mitre_id: str | None = None,
    origins: Iterable[str] = (),
    reference: dict[str, Any] | None = None,
) -> str:
    names = [clean_name(name), *(clean_name(x) for x in aliases)]
    names = list(dict.fromkeys(x for x in names if useful_name(x)))
    if not names:
        return ""
    overlapping = {
        actor_id
        for item in names
        for actor_id in alias_index.get(normalized_name(item), set())
    }
    actor_id = identity_id(mitre_id, names[0])
    if mitre_id:
        mitre_actor_id = f"actor-census--mitre:{mitre_id}"
        if mitre_actor_id in identities:
            actor_id = mitre_actor_id
        elif len(overlapping) == 1:
            old_id = next(iter(overlapping))
            old = identities.pop(old_id)
            old["actor_id"] = mitre_actor_id
            identities[mitre_actor_id] = old
            for ids in alias_index.values():
                if old_id in ids:
                    ids.remove(old_id)
                    ids.add(mitre_actor_id)
            actor_id = mitre_actor_id
    elif len(overlapping) == 1:
        actor_id = next(iter(overlapping))
    item = identities.setdefault(
        actor_id,
        {
            "actor_id": actor_id,
            "canonical_name": names[0],
            "aliases": [],
            "mitre_group_id": mitre_id,
            "origins": [],
            "reference_evidence": [],
            "mentions": [],
        },
    )
    if mitre_id and not item.get("mitre_group_id"):
        item["mitre_group_id"] = mitre_id
    item["aliases"] = list(dict.fromkeys([*item["aliases"], *names]))
    item["origins"] = list(dict.fromkeys([*item["origins"], *origins]))
    if reference and reference not in item["reference_evidence"]:
        item["reference_evidence"].append(reference)
    for alias in names:
        alias_index[normalized_name(alias)].add(actor_id)
    return actor_id


def load_microsoft_names(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    results: list[dict[str, Any]] = []
    try:
        sheet = workbook["alphabetical"]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [clean_name(str(value)) if value is not None else "" for value in row]
            if row_number <= 3:
                continue
            for offset in (1, 6):
                previous = values[offset] if offset < len(values) else ""
                current = values[offset + 1] if offset + 1 < len(values) else ""
                origin = values[offset + 2] if offset + 2 < len(values) else ""
                other = values[offset + 3] if offset + 3 < len(values) else ""
                if not current:
                    continue
                results.append(
                    {
                        "name": current,
                        "aliases": [previous, *split_alias_cell(other)],
                        "origin": origin,
                        "row": row_number,
                    }
                )
    finally:
        workbook.close()
    return results


def corpus_files(root: Path) -> list[Path]:
    files = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        relative = path.relative_to(root)
        if any(part in SKIP_DIRS for part in relative.parts):
            continue
        if relative.parts[:2] == ("actor_profile", "reference"):
            continue
        files.append(path)
    return sorted(files)


def build_alias_patterns(
    alias_index: dict[str, set[str]], identities: dict[str, Any]
) -> tuple[re.Pattern[str], re.Pattern[str] | None]:
    surface_names = {
        alias
        for item in identities.values()
        for alias in item["aliases"]
        if useful_name(alias)
    }
    insensitive = {
        name for name in surface_names
        if not (
            len(name.split()) == 1
            and name.casefold() in AMBIGUOUS_SINGLE_NAMES
        )
    }
    sensitive = surface_names - insensitive
    alternatives = sorted(
        (re.escape(name) for name in insensitive),
        key=len,
        reverse=True,
    )
    insensitive_pattern = re.compile(
        r"(?<![A-Za-z0-9])(?:" + "|".join(alternatives) + r")(?![A-Za-z0-9])",
        re.IGNORECASE,
    )
    sensitive_pattern = (
        re.compile(
            r"(?<![A-Za-z0-9])(?:"
            + "|".join(sorted((re.escape(name) for name in sensitive), key=len, reverse=True))
            + r")(?![A-Za-z0-9])"
        )
        if sensitive
        else None
    )
    return insensitive_pattern, sensitive_pattern


def context(text: str, start: int, end: int, width: int = 220) -> str:
    left = max(0, start - width)
    right = min(len(text), end + width)
    return " ".join(text[left:right].replace("\u0000", " ").split())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repository-root", type=Path, default=Path.cwd())
    parser.add_argument("--catalog", type=Path, default=Path("actor_profile/corpus-catalog.json"))
    parser.add_argument("--output", type=Path, default=Path("actor_profile/actor-census.json"))
    parser.add_argument("--csv-output", type=Path, default=Path("actor_profile/actor-census.csv"))
    parser.add_argument("--max-mentions-per-actor-source", type=int, default=5)
    args = parser.parse_args()

    root = args.repository_root.resolve()
    catalog = load_json((root / args.catalog).resolve())
    attack = load_json((root / catalog["reference_sources"]["mitre_attack_index"]).resolve())
    identities: dict[str, dict[str, Any]] = {}
    alias_index: dict[str, set[str]] = defaultdict(set)

    for mitre_id, group in attack["groups"].items():
        add_identity(
            identities,
            alias_index,
            name=group["name"],
            aliases=group.get("aliases", []),
            mitre_id=mitre_id,
            reference={"source": "MITRE ATT&CK", "external_id": mitre_id},
        )

    workbook_path = root / catalog["reference_sources"]["actor_mapping_workbook"]
    for record in load_workbook_rows(workbook_path):
        names = workbook_actor_names(record)
        if not names:
            continue
        add_identity(
            identities,
            alias_index,
            name=record["fields"].get("Common Name", names[0]),
            aliases=names,
            mitre_id=(
                record["fields"].get("MITRE ATT&CK")
                if re.fullmatch(r"G\d{4}", record["fields"].get("MITRE ATT&CK", ""))
                else None
            ),
            origins=[record["sheet"]],
            reference={
                "source": workbook_path.name,
                "sheet": record["sheet"],
                "row": record["row"],
            },
        )

    microsoft_path = root / catalog["reference_sources"]["microsoft_mapping_workbook"]
    for record in load_microsoft_names(microsoft_path):
        add_identity(
            identities,
            alias_index,
            name=record["name"],
            aliases=record["aliases"],
            origins=[record["origin"]] if record["origin"] else [],
            reference={"source": microsoft_path.name, "sheet": "alphabetical", "row": record["row"]},
        )

    catalog_aliases: set[str] = set()
    for actor in catalog["actors"]:
        names = [actor["name"], *actor.get("aliases", [])]
        catalog_aliases.update(normalized_name(name) for name in names)
        add_identity(
            identities,
            alias_index,
            name=actor["name"],
            aliases=actor.get("aliases", []),
            mitre_id=actor.get("mitre_group_id"),
            reference={"source": "corpus-catalog.json", "slug": actor["slug"]},
        )

    alias_pattern, sensitive_alias_pattern = build_alias_patterns(alias_index, identities)
    files = corpus_files(root)
    mention_keys: set[tuple[str, str, str, int, int]] = set()
    mention_counts: dict[tuple[str, str], int] = defaultdict(int)
    errors: list[dict[str, str]] = []

    for file_index, path in enumerate(files, start=1):
        relative = path.relative_to(root).as_posix()
        try:
            for record in source_records(path, path.suffix.lower()):
                text = record["text"]
                if not text:
                    continue
                found: list[tuple[str, int, int, str]] = []
                for match in alias_pattern.finditer(text):
                    normalized = normalized_name(match.group())
                    for actor_id in alias_index.get(normalized, set()):
                        found.append((actor_id, match.start(), match.end(), match.group()))
                if sensitive_alias_pattern:
                    for match in sensitive_alias_pattern.finditer(text):
                        normalized = normalized_name(match.group())
                        for actor_id in alias_index.get(normalized, set()):
                            found.append((actor_id, match.start(), match.end(), match.group()))
                for match in IDENTIFIER_RE.finditer(text):
                    surface = clean_name(match.group())
                    normalized = normalized_name(surface)
                    actor_ids = alias_index.get(normalized, set())
                    if not actor_ids:
                        actor_id = add_identity(
                            identities,
                            alias_index,
                            name=surface,
                            reference={"source": "corpus-pattern-discovery"},
                        )
                        actor_ids = {actor_id}
                    for actor_id in actor_ids:
                        found.append((actor_id, match.start(), match.end(), surface))
                location = json.dumps(record["location"], sort_keys=True)
                for actor_id, start, end, surface in found:
                    count_key = (actor_id, relative)
                    if mention_counts[count_key] >= args.max_mentions_per_actor_source:
                        continue
                    key = (actor_id, relative, location, start, end)
                    if key in mention_keys:
                        continue
                    mention_keys.add(key)
                    mention_counts[count_key] += 1
                    identities[actor_id]["mentions"].append(
                        {
                            "source_path": relative,
                            "source_location": record["location"],
                            "matched_name": surface,
                            "context_excerpt": context(text, start, end),
                        }
                    )
        except Exception as exc:
            errors.append(
                {"source_path": relative, "error": f"{type(exc).__name__}: {exc}"}
            )
        if file_index % 25 == 0 or file_index == len(files):
            print(
                json.dumps(
                    {
                        "processed_files": file_index,
                        "total_files": len(files),
                        "identities": len(identities),
                        "mentions": sum(len(x["mentions"]) for x in identities.values()),
                        "errors": len(errors),
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )

    catalog_by_norm = {
        normalized_name(name): actor["slug"]
        for actor in catalog["actors"]
        for name in [actor["name"], *actor.get("aliases", [])]
    }
    for item in identities.values():
        slugs = {
            catalog_by_norm[normalized_name(alias)]
            for alias in item["aliases"]
            if normalized_name(alias) in catalog_by_norm
        }
        item["catalog_slugs"] = sorted(slugs)
        item["is_profiled"] = bool(slugs)
        item["mention_source_count"] = len(
            {mention["source_path"] for mention in item["mentions"]}
        )
        item["mention_count"] = len(item["mentions"])
        item["aliases"] = sorted(set(item["aliases"]), key=str.casefold)
        item["mentions"].sort(
            key=lambda x: (
                x["source_path"],
                json.dumps(x["source_location"], sort_keys=True),
            )
        )

    actor_rows = sorted(
        identities.values(),
        key=lambda x: (not x["is_profiled"], x["canonical_name"].casefold()),
    )
    result = {
        "schema_version": "1.0.0",
        "generated_at": utc_now(),
        "scope": {
            "repository_root": str(root),
            "supported_suffixes": sorted(SUPPORTED_SUFFIXES),
            "file_count": len(files),
            "all_repository_roots_included": True,
        },
        "counts": {
            "actor_identities": len(actor_rows),
            "profiled": sum(x["is_profiled"] for x in actor_rows),
            "unprofiled": sum(not x["is_profiled"] for x in actor_rows),
            "with_corpus_mentions": sum(bool(x["mentions"]) for x in actor_rows),
            "mentions": sum(len(x["mentions"]) for x in actor_rows),
            "source_errors": len(errors),
        },
        "source_errors": errors,
        "actors": actor_rows,
    }
    write_json_atomic((root / args.output).resolve(), result)
    csv_path = (root / args.csv_output).resolve()
    with csv_path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=[
                "actor_id", "canonical_name", "mitre_group_id", "is_profiled",
                "catalog_slugs", "mention_source_count", "mention_count", "aliases",
            ],
        )
        writer.writeheader()
        for item in actor_rows:
            writer.writerow(
                {
                    "actor_id": item["actor_id"],
                    "canonical_name": item["canonical_name"],
                    "mitre_group_id": item.get("mitre_group_id") or "",
                    "is_profiled": str(item["is_profiled"]).lower(),
                    "catalog_slugs": json.dumps(item["catalog_slugs"], ensure_ascii=False),
                    "mention_source_count": item["mention_source_count"],
                    "mention_count": item["mention_count"],
                    "aliases": json.dumps(item["aliases"], ensure_ascii=False),
                }
            )
    print(json.dumps({"output": str(args.output), "csv": str(args.csv_output), **result["counts"]}))
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
