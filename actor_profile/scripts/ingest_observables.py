#!/usr/bin/env python3
"""Extract and normalize IOC and non-IOC artifacts from repository sources.

The script never opens archives or executes samples. Ambiguous values are retained
as candidates rather than silently discarded.
"""

from __future__ import annotations

import argparse
import csv
import ipaddress
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Iterator

from common import (
    earliest_time,
    json_array_cell,
    latest_time,
    load_json,
    merge_metadata,
    normalize_observable,
    normalize_time,
    refang,
    resolve_inside,
    stable_id,
    stix_pattern,
    unknown_time,
    utc_now,
    write_json_atomic,
)


SUPPORTED_SUFFIXES = {
    ".pdf",
    ".xlsx",
    ".csv",
    ".tsv",
    ".json",
    ".stix2",
    ".md",
    ".txt",
}

HASH_RE = re.compile(r"(?<![0-9A-Fa-f])(?:[0-9A-Fa-f][ :.-]?){32,128}(?![0-9A-Fa-f])")
IPV4_RE = re.compile(r"(?<!\d)(?:\d{1,3}(?:\.|\[\.\]|\(\.\))){3}\d{1,3}(?!\d)")
IPV6_RE = re.compile(r"(?<![0-9A-Fa-f:])(?:[0-9A-Fa-f]{1,4}:){2,7}[0-9A-Fa-f]{0,4}(?![0-9A-Fa-f:])")
URL_RE = re.compile(
    r"\b(?:https?|hxxps?)(?::|\[:\])//[^\s<>\"'`]+",
    re.IGNORECASE,
)
EMAIL_RE = re.compile(
    r"\b[A-Z0-9._%+-]+@(?:[A-Z0-9-]+(?:\.|\[\.\]))+[A-Z]{2,63}\b",
    re.IGNORECASE,
)
DEFANGED_DOMAIN_RE = re.compile(
    r"\b(?:[A-Z0-9-]+(?:\[\.\]|\(\.\))){1,10}[A-Z]{2,63}\b",
    re.IGNORECASE,
)
PLAIN_DOMAIN_RE = re.compile(
    r"\b(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,63}\b",
    re.IGNORECASE,
)
FILE_EXTENSIONS = {
    "apk", "aspx", "bat", "cfg", "chm", "cmd", "com", "conf", "dat",
    "dll", "doc", "docm", "docx", "elf", "enc", "exe", "hta", "hwp",
    "ini", "js", "jse", "jsp", "lnk", "msc", "pdb", "pdf", "php",
    "pif", "ppt", "pptm", "pptx", "ps1", "rar", "rtf", "scr", "sys",
    "tmp", "vbe", "vbs", "xls", "xlsm", "xlsx", "zip",
    # Below are extensions that are not delegated TLDs, so treating them as a
    # filename can never shadow a real domain. Entries such as .md (Moldova),
    # .py (Paraguay), .sh, .io, .zip and .mov are deliberately left out above
    # because they are both file extensions and valid TLDs.
    "asp", "bin", "bmp", "class", "crt", "css", "csv", "db", "drv", "gz",
    "htm", "html", "ico", "img", "inf", "iso", "jar", "json", "jpeg", "jpg",
    "log", "msi", "ocx", "pem", "png", "pyc", "pyd", "reg", "sql", "sqlite",
    "svg", "tgz", "txt", "war", "xml", "yaml", "yml",
}

# 出典レポート自身の参考リンク(ベンダーブログ、CERT、報道、リファレンス)は
# IOCではない。ポータルの横串検索で誤結合を招くため取り込まない。
REFERENCE_HOSTS_PATH = (
    Path(__file__).resolve().parents[1] / "reference" / "reference-hosts.json"
)


def _load_reference_data(key: str) -> frozenset[str]:
    try:
        with REFERENCE_HOSTS_PATH.open(encoding="utf-8") as handle:
            return frozenset(
                entry.strip().lower()
                for entry in json.load(handle).get(key, [])
                if entry.strip()
            )
    except (OSError, ValueError):
        return frozenset()


REFERENCE_HOSTS = _load_reference_data("hosts")
# co.kr や ddns.net のような公開サフィックスは、それ単体では指標にならない。
# サブドメイン (mfahost.ddns.net) は実際のIOCなので完全一致のときだけ弾く。
PUBLIC_SUFFIXES = _load_reference_data("public_suffixes")
# 公開DNSリゾルバ。技術的には到達可能だが指標にはならない。
PUBLIC_RESOLVERS = frozenset({
    "8.8.8.8", "8.8.4.4", "1.1.1.1", "1.0.0.1", "9.9.9.9", "149.112.112.112",
    "208.67.222.222", "208.67.220.220", "4.2.2.1", "4.2.2.2", "114.114.114.114",
    "223.5.5.5", "180.76.76.76", "77.88.8.8",
})
DEFANG_MARKER_RE = re.compile(r"\[\s*[.@:/]\s*\]|\(\s*[.@]\s*\)|\{\s*\.\s*\}|hxxp", re.IGNORECASE)
URL_HOST_RE = re.compile(r"^[A-Za-z][A-Za-z0-9+.\-]*://([^/?#:]+)")

REGISTRY_RE = re.compile(
    r"\b(?:HKEY_[A-Z_]+|HKLM|HKCU|HKCR|HKU|HKCC)\\[^\r\n\t,;\"']+",
    re.IGNORECASE,
)
PDB_RE = re.compile(r"\b[A-Z]:\\[^\r\n\"']+?\.pdb\b", re.IGNORECASE)
NAMED_PIPE_RE = re.compile(r"\\\\\.\\pipe\\[^\s\"']+", re.IGNORECASE)
WINDOWS_PATH_RE = re.compile(
    r"(?:\b[A-Z]:\\|%[A-Z0-9_]+%\\)[^\r\n\t\"'<>|]+",
    re.IGNORECASE,
)
UNIX_PATH_RE = re.compile(
    r"(?<![A-Za-z0-9])/(?:tmp|var|etc|opt|usr|home|root|dev|proc)/[^\s\"'<>]+"
)
FILE_NAME_RE = re.compile(
    r"(?<![A-Za-z0-9_$@(){}\[\].-])[A-Za-z0-9_$@(){}\[\].-]{1,120}\."
    r"(?:exe|dll|sys|scr|pif|bat|cmd|ps1|vbs|vbe|js|jse|hta|chm|msc|"
    r"lnk|hwp|docm?|docx|xlsm?|xlsx|pptm?|pptx|rtf|apk|elf|php|jsp|aspx|"
    r"dat|cfg|conf|ini|tmp|enc|pdb|pdf|rar|zip|7z)(?![A-Za-z0-9_-])",
    re.IGNORECASE,
)
COMMAND_MARKER_RE = re.compile(
    r"\b(?:powershell(?:\.exe)?|cmd(?:\.exe)?\s+/[ck]|mshta(?:\.exe)?|"
    r"rundll32(?:\.exe)?|regsvr32(?:\.exe)?|certutil(?:\.exe)?|"
    r"schtasks(?:\.exe)?|wmic(?:\.exe)?|curl(?:\.exe)?|wget|bash\s+-c|sh\s+-c)\b",
    re.IGNORECASE,
)
MUTEX_CONTEXT_RE = re.compile(r"\bmutex\b", re.IGNORECASE)
STRING_CONTEXT_RE = re.compile(
    r"(?:interesting|string|magic|marker|parameter|암호|문자열)\s*[:=]\s*([^\r\n]{2,300})",
    re.IGNORECASE,
)
DATE_RE = re.compile(
    r"(?<!\d)(20\d{2})[-/.](\d{1,2})(?:[-/.](\d{1,2}))?(?!\d)"
)
IOC_CONTEXT_RE = re.compile(
    r"\b(?:ioc|indicator(?:s)? of compromise|c2|c&c|command and control|"
    r"hash(?:es)?|domain(?:s)?|ip(?: address)?|infrastructure|침해지표|위협지표)\b",
    re.IGNORECASE,
)


def short_context(text: str, limit: int = 500) -> str:
    return re.sub(r"\s+", " ", text).strip()[:limit]


def source_records(path: Path, suffix: str) -> Iterator[dict[str, Any]]:
    """Yield text-bearing records with precise source locations."""
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("pypdf is required for PDF ingestion") from exc
        reader = PdfReader(path, strict=False)
        for page_number, page in enumerate(reader.pages, start=1):
            yield {
                "text": page.extract_text() or "",
                "location": {"page": page_number},
                "fields": {},
                "method": "pdf-text",
            }
        return

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for XLSX ingestion") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                header: list[str] | None = None
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    values = ["" if value is None else str(value) for value in row]
                    if not any(values):
                        continue
                    if header is None:
                        header = [value.strip() for value in values]
                    fields = {
                        header[index]: value
                        for index, value in enumerate(values)
                        if header and index < len(header) and header[index] and value
                    }
                    yield {
                        "text": "\t".join(value for value in values if value),
                        "location": {"sheet": sheet.title, "row": row_number},
                        "fields": fields,
                        "method": "xlsx-row",
                    }
        finally:
            workbook.close()
        return

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
            reader = csv.DictReader(stream, delimiter=delimiter)
            if reader.fieldnames:
                for row_number, row in enumerate(reader, start=2):
                    fields = {str(k): str(v) for k, v in row.items() if k and v}
                    yield {
                        "text": "\t".join(fields.values()),
                        "location": {"row": row_number},
                        "fields": fields,
                        "method": f"{suffix[1:]}-row",
                    }
            else:
                stream.seek(0)
                for row_number, row in enumerate(
                    csv.reader(stream, delimiter=delimiter), start=1
                ):
                    yield {
                        "text": "\t".join(row),
                        "location": {"row": row_number},
                        "fields": {},
                        "method": f"{suffix[1:]}-row",
                    }
        return

    if suffix in {".json", ".stix2"}:
        data = load_json(path)

        def walk(value: Any, json_path: str = "$") -> Iterator[dict[str, Any]]:
            if isinstance(value, dict):
                if value.get("type") == "indicator" and isinstance(value.get("pattern"), str):
                    yield {
                        "text": value["pattern"],
                        "location": {"json_path": f"{json_path}.pattern"},
                        "fields": {
                            "stix_pattern": value["pattern"],
                            "valid_from": str(value.get("valid_from", "")),
                            "name": str(value.get("name", "")),
                        },
                        "method": "stix-indicator",
                    }
                for key, child in value.items():
                    yield from walk(child, f"{json_path}.{key}")
            elif isinstance(value, list):
                for index, child in enumerate(value):
                    yield from walk(child, f"{json_path}[{index}]")
            elif isinstance(value, str):
                yield {
                    "text": value,
                    "location": {"json_path": json_path},
                    "fields": {},
                    "method": "json-string",
                }

        yield from walk(data)
        return

    with path.open("r", encoding="utf-8", errors="replace") as stream:
        for line_number, line in enumerate(stream, start=1):
            if line.strip():
                yield {
                    "text": line.rstrip("\n"),
                    "location": {"line": line_number},
                    "fields": {},
                    "method": "text-line",
                }


def time_from_record(
    record: dict[str, Any], metadata: dict[str, Any]
) -> dict[str, Any]:
    field_map = metadata.get("field_map", {})
    observed_column = field_map.get("observed_at")
    if observed_column and record["fields"].get(observed_column):
        return normalize_time(
            record["fields"][observed_column], basis=f"column:{observed_column}"
        )
    match = DATE_RE.search(record["text"])
    if match:
        year, month, day = match.groups()
        if day:
            return normalize_time(f"{year}-{month}-{day}", basis="same-record")
        if month:
            return normalize_time(f"{year}-{month}", basis="same-record")
        return normalize_time(year, basis="same-record")
    default = metadata.get("default_observed_at", unknown_time())
    return default if default.get("status") != "unknown" else unknown_time()


def structured_metadata(
    record: dict[str, Any], metadata: dict[str, Any]
) -> dict[str, Any]:
    result = {
        "campaign_refs": list(metadata.get("campaign_refs", [])),
        "malware_refs": list(metadata.get("malware_refs", [])),
        "infrastructure_refs": list(metadata.get("infrastructure_refs", [])),
        "roles": list(metadata.get("roles", [])),
    }
    field_map = metadata.get("field_map", {})
    for target in result:
        column = field_map.get(target)
        if column and record["fields"].get(column):
            values = re.split(r"[,;|]", record["fields"][column])
            result[target] = sorted(
                set(result[target]) | {value.strip() for value in values if value.strip()}
            )
    return result


def classify_hash(raw: str) -> tuple[str, str] | None:
    compact = re.sub(r"[^0-9A-Fa-f]", "", raw)
    kind_by_length = {32: "md5", 40: "sha1", 64: "sha256", 128: "sha512"}
    kind = kind_by_length.get(len(compact))
    return (kind, compact) if kind else None


def analyst_marked(raw: str, explicit_structured: bool) -> bool:
    """出典側が指標として明示したか。

    難読化(``hxxp``、``[.]``)はアナリストが悪性だと判断した印であり、構造化IOC表
    からの取り込みも同様。どちらかに当てはまるなら参考ホスト判定より優先する。
    """
    return explicit_structured or bool(DEFANG_MARKER_RE.search(raw))


def routable_address(address: ipaddress.IPv4Address | ipaddress.IPv6Address) -> bool:
    """指標になり得るアドレスかどうか。

    ループバック、RFC1918、ドキュメント用、マルチキャスト等は誰の環境にも現れるため
    指標にならない。公開DNSリゾルバも同様に横串で誤結合するだけなので除外する。
    """
    if (
        address.is_private
        or address.is_loopback
        or address.is_reserved
        or address.is_multicast
        or address.is_unspecified
        or address.is_link_local
    ):
        return False
    return str(address) not in PUBLIC_RESOLVERS


def host_of(raw: str) -> str:
    """URL・ドメイン・メールアドレスからホスト部分を取り出す。"""
    value = refang(raw).strip().strip("<>\"'`")
    match = URL_HOST_RE.match(value)
    host = (match.group(1) if match else value).lower().rstrip(".")
    if "@" in host:
        host = host.rsplit("@", 1)[-1]
    return host[4:] if host.startswith("www.") else host


def reference_host(raw: str) -> bool:
    """値のホストが出典レポートの参考リンク側かどうか。"""
    if not REFERENCE_HOSTS:
        return False
    host = host_of(raw)
    if host in REFERENCE_HOSTS:
        return True
    # ドット区切りのサフィックス一致(blog.securelist.com -> securelist.com)
    labels = host.split(".")
    return any(
        ".".join(labels[i:]) in REFERENCE_HOSTS for i in range(1, len(labels) - 1)
    )


def plausible_domain(raw: str) -> bool:
    """Reject common filename/OCR shapes before treating text as a domain."""
    value = refang(raw).strip().strip(".").lower()
    labels = value.split(".")
    if len(labels) < 2 or labels[-1] in FILE_EXTENSIONS:
        return False
    # co.kr や ddns.net といった公開サフィックスそのものは指標にならない。
    if value in PUBLIC_SUFFIXES:
        return False
    # Masked IPs such as 192.0.2.xxx are artifacts, not valid domain IOCs.
    if len(labels) == 4 and all(label.isdigit() for label in labels[:3]):
        return False
    return True


def extract_iocs(
    text: str, *, allow_plain_domains: bool, explicit_structured: bool
) -> list[tuple[str, str, str]]:
    """Return kind, raw value, disposition."""
    results: list[tuple[str, str, str]] = []
    context_ioc = bool(IOC_CONTEXT_RE.search(text))
    defanged = "[.]" in text or "(.)" in text or "hxxp" in text.lower()
    disposition = "confirmed" if (context_ioc or defanged or explicit_structured) else "candidate"

    occupied: list[tuple[int, int]] = []
    for match in HASH_RE.finditer(text):
        classified = classify_hash(match.group())
        if classified:
            kind, raw = classified
            results.append((kind, raw, disposition))
            occupied.append(match.span())

    for match in URL_RE.finditer(text):
        raw = match.group()
        occupied.append(match.span())
        if not analyst_marked(raw, explicit_structured) and reference_host(raw):
            continue  # 出典レポートの参考リンク
        results.append(("url", raw, disposition))
    for match in EMAIL_RE.finditer(text):
        raw = match.group()
        occupied.append(match.span())
        if not analyst_marked(raw, explicit_structured) and reference_host(raw):
            continue  # ベンダーの問い合わせ窓口など、出典側の連絡先
        results.append(("email", raw, disposition))
    for match in IPV4_RE.finditer(text):
        raw = refang(match.group())
        try:
            address = ipaddress.IPv4Address(raw)
        except ipaddress.AddressValueError:
            continue
        occupied.append(match.span())
        if not routable_address(address):
            continue
        results.append(("ipv4", match.group(), disposition))
    for match in IPV6_RE.finditer(text):
        try:
            address = ipaddress.IPv6Address(match.group())
        except ipaddress.AddressValueError:
            continue
        occupied.append(match.span())
        if not routable_address(address):
            continue
        results.append(("ipv6", match.group(), disposition))
    for match in DEFANGED_DOMAIN_RE.finditer(text):
        if plausible_domain(match.group()):
            results.append(("domain", match.group(), disposition))
            occupied.append(match.span())
    if allow_plain_domains or context_ioc or explicit_structured:
        for match in PLAIN_DOMAIN_RE.finditer(text):
            if any(start <= match.start() < end for start, end in occupied):
                continue
            raw = match.group()
            if not plausible_domain(raw):
                continue
            if not analyst_marked(raw, explicit_structured) and reference_host(raw):
                continue  # 出典レポートの参考リンク
            results.append(("domain", raw, disposition))
    return results


def artifact_match(
    kind: str, raw: str, disposition: str, results: list[tuple[str, str, str]]
) -> None:
    value = raw.strip().strip("`'\"")
    if value:
        results.append((kind, value, disposition))


def extract_artifacts(
    text: str, *, explicit_structured: bool
) -> list[tuple[str, str, str]]:
    results: list[tuple[str, str, str]] = []
    # Mere occurrence in an IOC appendix does not prove that prose mentioning a
    # command/path is an actor-specific artifact. Structured mappings may confirm;
    # heuristic PDF/text extraction remains reviewable.
    disposition = "confirmed" if explicit_structured else "candidate"

    for match in NAMED_PIPE_RE.finditer(text):
        artifact_match("named-pipe", match.group(), disposition, results)
    for match in PDB_RE.finditer(text):
        artifact_match("pdb-path", match.group(), disposition, results)
    for match in REGISTRY_RE.finditer(text):
        artifact_match("registry-key", match.group(), disposition, results)
    for match in WINDOWS_PATH_RE.finditer(text):
        value = match.group()
        if value.lower().endswith(".pdb"):
            continue
        artifact_match("file-path", value, disposition, results)
    for match in UNIX_PATH_RE.finditer(text):
        artifact_match("file-path", match.group(), disposition, results)
    for match in FILE_NAME_RE.finditer(text):
        value = match.group().strip()
        if "\\" in value or "/" in value:
            continue
        artifact_match("file-name", value, disposition, results)
    for line in text.splitlines():
        if COMMAND_MARKER_RE.search(line):
            artifact_match("command", short_context(line, 1000), disposition, results)
    string_match = STRING_CONTEXT_RE.search(text)
    if string_match:
        artifact_match("sample-string", string_match.group(1), disposition, results)
    if MUTEX_CONTEXT_RE.search(text):
        match = re.search(r"mutex\s*[:=]\s*([^\s,;]{2,200})", text, re.IGNORECASE)
        if match:
            artifact_match("mutex", match.group(1), disposition, results)
    return results


def explicit_mapped_values(
    record: dict[str, Any], metadata: dict[str, Any]
) -> tuple[list[tuple[str, str, str]], list[tuple[str, str, str]]]:
    field_map = metadata.get("field_map", {})
    fields = record["fields"]
    iocs: list[tuple[str, str, str]] = []
    artifacts: list[tuple[str, str, str]] = []
    value_column = field_map.get("value")
    type_column = field_map.get("type")
    artifact_type_column = field_map.get("artifact_type")
    if value_column and fields.get(value_column):
        raw = fields[value_column]
        declared_type = fields.get(type_column, "").lower() if type_column else ""
        if declared_type in {
            "md5", "sha1", "sha256", "sha512", "ipv4", "ipv6",
            "domain", "url", "email", "certificate-fingerprint",
        }:
            iocs.append((declared_type, raw, "confirmed"))
        elif artifact_type_column and fields.get(artifact_type_column):
            artifacts.append((fields[artifact_type_column], raw, "confirmed"))
    return iocs, artifacts


def expand_sources(
    manifest: dict[str, Any], repository_root: Path
) -> list[dict[str, Any]]:
    defaults = manifest["defaults"]
    expanded: list[dict[str, Any]] = []
    seen: set[Path] = set()
    for item in manifest.get("sources", []):
        metadata = merge_metadata(defaults, item)
        path = resolve_inside(repository_root, item["path"])
        if path in seen:
            continue
        seen.add(path)
        metadata["resolved_path"] = path
        expanded.append(metadata)
    for group in manifest.get("source_groups", []):
        metadata = merge_metadata(defaults, group)
        for path in sorted(repository_root.glob(group["path_glob"])):
            if not path.is_file() or path.suffix.lower() not in SUPPORTED_SUFFIXES:
                continue
            path = path.resolve()
            if path in seen:
                continue
            seen.add(path)
            item = dict(metadata)
            relative = path.relative_to(repository_root).as_posix()
            item["source_id"] = (
                f"{group['source_id_prefix']}--"
                f"{stable_id('file', relative).split(':', 1)[1][:16]}"
            )
            item["path"] = relative
            item["resolved_path"] = path
            expanded.append(item)
    return expanded


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("manifest", type=Path)
    parser.add_argument("--iocs-output", type=Path)
    parser.add_argument("--artifacts-output", type=Path)
    args = parser.parse_args()

    manifest_path = args.manifest.resolve()
    manifest = load_json(manifest_path)
    repository_root_value = Path(manifest.get("repository_root", "../.."))
    repository_root = (
        repository_root_value
        if repository_root_value.is_absolute()
        else manifest_path.parent / repository_root_value
    ).resolve()
    if not repository_root.is_dir():
        raise ValueError(f"Repository root is not a directory: {repository_root}")
    actor_ref = manifest["actor_ref"]
    iocs_output = (
        args.iocs_output.resolve()
        if args.iocs_output
        else manifest_path.parent / "iocs.json"
    )
    artifacts_output = (
        args.artifacts_output.resolve()
        if args.artifacts_output
        else manifest_path.parent / "artifacts.csv"
    )

    source_items = expand_sources(manifest, repository_root)
    indicator_observations: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    indicator_raw: dict[tuple[str, str], str] = {}
    indicator_dispositions: dict[tuple[str, str], set[str]] = defaultdict(set)
    artifact_rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    processed = 0

    for source in source_items:
        path = source["resolved_path"]
        if not path.exists():
            errors.append({"source_id": source["source_id"], "error": "missing", "path": str(path)})
            continue
        suffix = source.get("format") or path.suffix.lower()
        if suffix and not str(suffix).startswith("."):
            suffix = f".{suffix}"
        try:
            for record in source_records(path, str(suffix).lower()):
                text = record["text"]
                if not text.strip():
                    continue
                mapped_iocs, mapped_artifacts = explicit_mapped_values(record, source)
                explicit = bool(mapped_iocs or mapped_artifacts or record["method"] == "stix-indicator")
                ioc_matches = mapped_iocs + extract_iocs(
                    text,
                    allow_plain_domains=bool(source.get("allow_plain_domains", False)),
                    explicit_structured=explicit,
                )
                artifact_matches = mapped_artifacts + extract_artifacts(
                    text, explicit_structured=explicit
                )
                observed_at = time_from_record(record, source)
                source_published = normalize_time(
                    source.get("published_at"), basis="source-publication"
                )
                related = structured_metadata(record, source)
                common = {
                    "observed_at": observed_at,
                    "source_published_at": source_published,
                    "source_id": source["source_id"],
                    "source_path": path.relative_to(repository_root).as_posix(),
                    "source_location": record["location"],
                    **related,
                    "confidence": source.get("confidence", "medium"),
                    "tlp": source.get("tlp", "TLP:CLEAR"),
                    "extraction_method": record["method"],
                    "context_excerpt": short_context(text),
                    "analyst_notes": source.get("analyst_notes", ""),
                }
                for kind, raw, disposition in ioc_matches:
                    normalized = normalize_observable(kind, raw)
                    if not normalized:
                        continue
                    key = (kind, normalized)
                    location_json = json.dumps(record["location"], sort_keys=True)
                    observation = {
                        "observation_id": stable_id(
                            "observation", actor_ref, kind, normalized,
                            source["source_id"], location_json
                        ),
                        **common,
                        "raw_value": raw,
                    }
                    if observation["observation_id"] not in {
                        item["observation_id"] for item in indicator_observations[key]
                    }:
                        indicator_observations[key].append(observation)
                    indicator_raw.setdefault(key, refang(raw))
                    indicator_dispositions[key].add(disposition)
                for artifact_type, raw, disposition in artifact_matches:
                    normalized = normalize_observable("artifact", raw)
                    if not normalized:
                        continue
                    artifact_id = stable_id("artifact", artifact_type, normalized)
                    location_json = json.dumps(record["location"], sort_keys=True)
                    artifact_rows.append(
                        {
                            "schema_version": "1.0.0",
                            "actor_ref": actor_ref,
                            "artifact_id": artifact_id,
                            "observation_id": stable_id(
                                "observation", actor_ref, artifact_type, normalized,
                                source["source_id"], location_json
                            ),
                            "artifact_type": artifact_type,
                            "value": raw,
                            "normalized_value": normalized,
                            "disposition": disposition,
                            "observed_at": observed_at.get("value") or "",
                            "observed_at_precision": observed_at["precision"],
                            "observed_at_status": observed_at["status"],
                            "observed_at_basis": observed_at["basis"],
                            "source_published_at": source_published.get("value") or "",
                            "source_id": source["source_id"],
                            "source_path": common["source_path"],
                            "source_location": json.dumps(record["location"], ensure_ascii=False, sort_keys=True),
                            "campaign_refs": json_array_cell(related["campaign_refs"]),
                            "malware_refs": json_array_cell(related["malware_refs"]),
                            "infrastructure_refs": json_array_cell(related["infrastructure_refs"]),
                            "roles": json_array_cell(related["roles"]),
                            "campaign_count": 0,
                            "seen_in_multiple_campaigns": "false",
                            "confidence": common["confidence"],
                            "tlp": common["tlp"],
                            "extraction_method": record["method"],
                            "context_excerpt": common["context_excerpt"],
                            "analyst_notes": common["analyst_notes"],
                        }
                    )
            processed += 1
        except Exception as exc:
            errors.append(
                {
                    "source_id": source["source_id"],
                    "path": str(path),
                    "error": f"{type(exc).__name__}: {exc}",
                }
            )

    indicators = []
    for (kind, normalized), observations in sorted(indicator_observations.items()):
        campaigns = sorted(
            {ref for observation in observations for ref in observation["campaign_refs"]}
        )
        malware = sorted(
            {ref for observation in observations for ref in observation["malware_refs"]}
        )
        infrastructure = sorted(
            {ref for observation in observations for ref in observation["infrastructure_refs"]}
        )
        roles = sorted(
            {role for observation in observations for role in observation["roles"]}
        )
        dispositions = indicator_dispositions[(kind, normalized)]
        disposition = (
            "confirmed" if "confirmed" in dispositions
            else "candidate" if "candidate" in dispositions
            else "rejected"
        )
        indicators.append(
            {
                "indicator_id": stable_id("indicator", kind, normalized),
                "type": kind,
                "value": indicator_raw[(kind, normalized)],
                "normalized_value": normalized,
                "stix_pattern": stix_pattern(kind, normalized),
                "disposition": disposition,
                "first_observed": earliest_time(o["observed_at"] for o in observations),
                "last_observed": latest_time(o["observed_at"] for o in observations),
                "observation_count": len(observations),
                "campaign_count": len(campaigns),
                "seen_in_multiple_campaigns": len(campaigns) > 1,
                "campaign_refs": campaigns,
                "malware_refs": malware,
                "infrastructure_refs": infrastructure,
                "roles": roles,
                "observations": sorted(
                    observations,
                    key=lambda item: (
                        item["source_path"],
                        json.dumps(item["source_location"], sort_keys=True),
                    ),
                ),
            }
        )

    # Deduplicate exact artifact observations, then annotate cross-campaign reuse.
    artifact_by_observation = {
        row["observation_id"]: row for row in artifact_rows
    }
    artifact_rows = list(artifact_by_observation.values())
    campaigns_by_artifact: dict[str, set[str]] = defaultdict(set)
    for row in artifact_rows:
        campaigns_by_artifact[row["artifact_id"]].update(
            json.loads(row["campaign_refs"])
        )
    for row in artifact_rows:
        count = len(campaigns_by_artifact[row["artifact_id"]])
        row["campaign_count"] = count
        row["seen_in_multiple_campaigns"] = "true" if count > 1 else "false"
    artifact_rows.sort(
        key=lambda row: (
            row["artifact_type"],
            row["normalized_value"],
            row["source_path"],
            row["source_location"],
        )
    )

    dataset = {
        "schema_version": "1.0.0",
        "actor_ref": actor_ref,
        "generated_at": utc_now(),
        "sources": [
            {
                "source_id": item["source_id"],
                "path": item["resolved_path"].relative_to(repository_root).as_posix(),
                "published_at": normalize_time(
                    item.get("published_at"), basis="source-publication"
                ),
                "confidence": item.get("confidence", "medium"),
                "tlp": item.get("tlp", "TLP:CLEAR"),
                "analyst_notes": item.get("analyst_notes", ""),
            }
            for item in source_items
        ],
        "indicators": indicators,
        "ingestion": {
            "source_count": len(source_items),
            "processed_source_count": processed,
            "error_source_count": len(errors),
            "candidate_count": sum(
                indicator["disposition"] == "candidate" for indicator in indicators
            ),
            "errors": errors,
        },
    }
    write_json_atomic(iocs_output, dataset)

    columns_spec = load_json(
        Path(__file__).resolve().parent.parent
        / "schemas"
        / "artifacts-csv-columns.json"
    )
    artifacts_output.parent.mkdir(parents=True, exist_ok=True)
    with artifacts_output.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(
            stream, fieldnames=columns_spec["columns"], extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(artifact_rows)

    print(
        json.dumps(
            {
                "actor_ref": actor_ref,
                "sources": len(source_items),
                "processed": processed,
                "errors": len(errors),
                "indicators": len(indicators),
                "indicator_observations": sum(
                    item["observation_count"] for item in indicators
                ),
                "artifacts": len(artifact_rows),
                "iocs_output": str(iocs_output),
                "artifacts_output": str(artifacts_output),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
