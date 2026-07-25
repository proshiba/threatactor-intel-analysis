#!/usr/bin/env python3
"""Bootstrap standardized profiles for every actor in corpus-catalog.json.

The script combines three local evidence layers:
1. actor-specific repository documents,
2. the compact official MITRE ATT&CK index, and
3. the repository's actor-mapping workbook.

Existing profiles are preserved unless --overwrite is specified.
"""

from __future__ import annotations

import argparse
import copy
import json
import re
from pathlib import Path
from typing import Any, Iterable

from common import (
    load_json,
    normalize_time,
    slugify,
    stable_digest,
    unknown_time,
    utc_now,
    write_json_atomic,
)
from ingest_observables import SUPPORTED_SUFFIXES, source_records


SCRIPT_DIR = Path(__file__).resolve().parent
FRAMEWORK_DIR = SCRIPT_DIR.parent
TECHNIQUE_RE = re.compile(r"(?<![A-Z0-9])T\d{4}(?:\.\d{3})?(?!\d)")
DATE_IN_NAME_RE = re.compile(r"(?<!\d)(20\d{2})[-_]?(\d{2})?[-_]?(\d{2})?(?!\d)")

COUNTRIES = {
    "afghanistan", "albania", "argentina", "armenia", "australia", "austria",
    "azerbaijan", "bangladesh", "belarus", "belgium", "brazil", "canada",
    "china", "colombia", "czech republic", "ecuador", "egypt", "estonia",
    "france", "georgia", "germany", "hong kong", "india", "indonesia",
    "iran", "iraq", "israel", "italy", "japan", "jordan", "kazakhstan",
    "korea", "kyrgyzstan", "latvia", "lebanon", "lithuania", "malaysia",
    "mexico", "mongolia", "netherlands", "north korea", "norway", "pakistan",
    "philippines", "poland", "romania", "russia", "saudi arabia",
    "singapore", "south korea", "spain", "sri lanka", "sweden", "switzerland",
    "syria", "taiwan", "tajikistan", "thailand", "turkey", "ukraine",
    "united arab emirates", "united kingdom", "united states", "uzbekistan",
    "vietnam",
}

SECTOR_PATTERNS = {
    "Government": r"\bgovernment|ministr(?:y|ies)|public sector|diplomat",
    "Defense": r"\bdefen[cs]e|military|armed forces|aerospace",
    "Technology": r"\btechnology|high tech|software|information technology|\bit\b",
    "Telecommunications": r"\btelecom|telecommunication",
    "Energy": r"\benergy|oil and gas|petroleum|electric",
    "Finance": r"\bfinancial|finance|bank|cryptocurrency|insurance",
    "Healthcare": r"\bhealth|medical|pharmaceutical|biotech",
    "Education and Research": r"\beducation|university|academic|research",
    "Media": r"\bmedia|journalist|news",
    "Transportation": r"\btransport|aviation|shipping|maritime",
    "Manufacturing": r"\bmanufactur|industrial|engineering",
    "Retail and Hospitality": r"\bretail|hotel|hospitality|restaurant",
    "Nonprofit and Civil Society": r"\bngo|nonprofit|civil society|activist",
}


def normalized_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def unique(values: Iterable[str]) -> list[str]:
    return list(dict.fromkeys(value.strip() for value in values if value and value.strip()))


def source_id(actor_slug: str, relative_path: str) -> str:
    digest = stable_digest(actor_slug, relative_path)[:16]
    return f"source--{actor_slug}--{digest}"


def title_from_path(path: Path) -> str:
    return re.sub(r"[_-]+", " ", path.stem).strip()


def date_from_filename(path: Path) -> dict[str, Any]:
    match = DATE_IN_NAME_RE.search(path.name)
    if not match:
        return unknown_time()
    year, month, day = match.groups()
    if month and day:
        return normalize_time(f"{year}-{month}-{day}", basis="filename")
    if month:
        return normalize_time(f"{year}-{month}", basis="filename")
    return normalize_time(year, basis="filename")


def source_type(path: Path) -> str:
    return {
        ".pdf": "report",
        ".xlsx": "spreadsheet",
        ".csv": "structured-data",
        ".tsv": "structured-data",
        ".json": "structured-data",
        ".stix2": "stix",
        ".md": "repository-notes",
        ".txt": "text-data",
    }.get(path.suffix.lower(), "document")


def expand_actor_files(root: Path, entries: list[str]) -> list[Path]:
    files: set[Path] = set()
    for entry in entries:
        candidate = (root / entry).resolve()
        try:
            candidate.relative_to(root)
        except ValueError as exc:
            raise ValueError(f"catalog path escapes repository: {entry}") from exc
        if candidate.is_file():
            if candidate.suffix.lower() in SUPPORTED_SUFFIXES:
                files.add(candidate)
            continue
        if candidate.is_dir():
            for path in candidate.rglob("*"):
                if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES:
                    files.add(path.resolve())
            continue
        raise FileNotFoundError(f"catalog source path does not exist: {entry}")
    return sorted(files)


def load_workbook_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for actor workbook enrichment") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            if sheet.title == "README" or sheet.title.startswith("_"):
                continue
            rows = list(sheet.iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, row in enumerate(rows)
                    if any(str(cell).strip() == "Common Name" for cell in row if cell)
                ),
                None,
            )
            if header_index is None:
                continue
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
            for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                values = [str(cell).strip() if cell is not None else "" for cell in row]
                fields = {
                    header: values[index]
                    for index, header in enumerate(headers)
                    if header and index < len(values) and values[index]
                }
                if fields.get("Common Name"):
                    records.append(
                        {
                            "sheet": sheet.title,
                            "row": row_number,
                            "headers": headers,
                            "values": values,
                            "fields": fields,
                        }
                    )
    finally:
        workbook.close()
    return records


def actor_name_cells(record: dict[str, Any]) -> list[str]:
    result = []
    for header, value in zip(record["headers"], record["values"]):
        if not header:
            continue
        lower = header.lower()
        if not value:
            continue
        if any(
            marker in lower
            for marker in (
                "operation", "toolset", "malware", "targets", "modus",
                "comment", "link",
            )
        ):
            continue
        if lower == "mitre att&ck":
            continue
        result.extend(
            item.strip()
            for item in re.split(r"\s+/\s+", value)
            if item.strip()
        )
    return result


def find_workbook_record(
    actor: dict[str, Any], records: list[dict[str, Any]]
) -> dict[str, Any] | None:
    mitre_id = actor.get("mitre_group_id")
    if mitre_id:
        for record in records:
            if record["fields"].get("MITRE ATT&CK", "").strip() == mitre_id:
                return record
    candidates = {
        normalized_name(actor["name"]),
        *(normalized_name(alias) for alias in actor.get("aliases", [])),
    }
    for record in records:
        names = {normalized_name(value) for value in actor_name_cells(record)}
        if candidates & names:
            return record
    return None


def workbook_operations(record: dict[str, Any] | None) -> list[str]:
    if not record:
        return []
    return unique(
        value
        for header, value in record["fields"].items()
        if header.lower().startswith("operation")
    )


def workbook_software(record: dict[str, Any] | None) -> list[str]:
    if not record:
        return []
    raw = record["fields"].get("Toolset / Malware", "")
    return unique(re.split(r"[,;\n]", raw))


def workbook_target_text(record: dict[str, Any] | None) -> str:
    return record["fields"].get("Targets", "") if record else ""


def time_point(value: Any, basis: str) -> dict[str, Any]:
    return normalize_time(value, basis=basis) if value else unknown_time()


def capability(
    prefix: str,
    name: str,
    description: str,
    evidence_refs: list[str],
    *,
    aliases: list[str] | None = None,
    types: list[str] | None = None,
    confidence: str = "medium",
) -> dict[str, Any]:
    return {
        "id": f"{prefix}--{slugify(name)}",
        "name": name,
        "aliases": aliases or [],
        "types": types or [],
        "description": description,
        "first_observed": unknown_time(),
        "last_observed": unknown_time(),
        "confidence": confidence,
        "evidence_refs": evidence_refs,
        "analyst_notes": "",
    }


def target(
    kind: str, name: str, description: str, evidence_refs: list[str]
) -> dict[str, Any]:
    return {
        "id": f"target--{slugify(kind)}--{slugify(name)}",
        "name": name,
        "description": description,
        "first_observed": unknown_time(),
        "last_observed": unknown_time(),
        "confidence": "medium",
        "evidence_refs": evidence_refs,
        "analyst_notes": "Automatically structured from target text; review scope and granularity.",
    }


def derive_targets(
    text: str, evidence_refs: list[str]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    lower = text.lower()
    countries = [
        target("country", name.title(), f"Targeting text mentions {name}.", evidence_refs)
        for name in sorted(COUNTRIES, key=len, reverse=True)
        if re.search(rf"\b{re.escape(name)}\b", lower)
    ]
    # Avoid retaining both "Korea" and a more specific Korea label.
    if any(item["name"] in {"North Korea", "South Korea"} for item in countries):
        countries = [item for item in countries if item["name"] != "Korea"]
    sectors = [
        target("sector", name, f"Targeting text indicates the {name} sector.", evidence_refs)
        for name, pattern in SECTOR_PATTERNS.items()
        if re.search(pattern, lower, re.IGNORECASE)
    ]
    return countries, sectors


def report_techniques(
    files: list[Path], root: Path, source_ids: dict[str, str]
) -> dict[str, dict[str, Any]]:
    results: dict[str, dict[str, Any]] = {}
    for path in files:
        suffix = path.suffix.lower()
        try:
            for record in source_records(path, suffix):
                text = record.get("text", "")
                for match in TECHNIQUE_RE.finditer(text):
                    technique_id = match.group()
                    result = results.setdefault(
                        technique_id, {"evidence_refs": [], "context": ""}
                    )
                    relative = path.relative_to(root).as_posix()
                    result["evidence_refs"].append(source_ids[relative])
                    if not result["context"]:
                        start = max(0, match.start() - 160)
                        end = min(len(text), match.end() + 280)
                        result["context"] = re.sub(
                            r"\s+", " ", text[start:end]
                        ).strip()
        except Exception:
            # Ingestion will separately report exact source errors. Profile
            # bootstrapping must continue and record the collection gap.
            continue
    for result in results.values():
        result["evidence_refs"] = sorted(set(result["evidence_refs"]))
    return results


def create_profile(
    actor: dict[str, Any],
    root: Path,
    files: list[Path],
    attack: dict[str, Any],
    workbook_record: dict[str, Any] | None,
    *,
    scan_report_ttps: bool,
) -> tuple[dict[str, Any], dict[str, Any]]:
    template = load_json(FRAMEWORK_DIR / "templates" / "actor-profile.template.json")
    now = utc_now()
    slug = actor["slug"]
    profile = copy.deepcopy(template)
    replacements = {
        "__ACTOR_NAME__": actor["name"],
        "__ACTOR_SLUG__": slug,
        "__NOW__": now,
    }

    def replace(value: Any) -> Any:
        if isinstance(value, str):
            for old, new in replacements.items():
                value = value.replace(old, new)
            return value
        if isinstance(value, list):
            return [replace(item) for item in value]
        if isinstance(value, dict):
            return {key: replace(item) for key, item in value.items()}
        return value

    profile = replace(profile)
    profile["actor"]["actor_types"] = actor.get("actor_types", [])

    reference_source_id = "source--mitre-attack-19-1"
    workbook_source_id = "source--actor-mapping-workbook"
    all_sources: list[dict[str, Any]] = [
        {
            "source_id": reference_source_id,
            "path": "actor_profile/reference/attack-index.json",
            "title": "MITRE Enterprise ATT&CK 19.1 compact local index",
            "publisher": "MITRE",
            "published_at": normalize_time("2026-05-12", basis="upstream-release"),
            "language": "en",
            "source_type": "structured-knowledge-base",
            "tlp": "TLP:CLEAR",
            "reliability": "high",
            "sha256": None,
            "analyst_notes": "Derived from the official MITRE attack-stix-data repository.",
        },
        {
            "source_id": workbook_source_id,
            "path": "APT Groups and Operations.xlsx",
            "title": "APT Groups and Operations",
            "publisher": "Florian Roth and community contributors",
            "published_at": unknown_time(),
            "language": "en",
            "source_type": "community-actor-mapping",
            "tlp": "TLP:CLEAR",
            "reliability": "medium",
            "sha256": None,
            "analyst_notes": "Mappings are leads, not definitive attribution; workbook disclaimer applies.",
        },
    ]
    source_ids: dict[str, str] = {}
    ingest_source_paths: set[str] = set()
    for path in files:
        relative = path.relative_to(root).as_posix()
        ingest_source_paths.add(relative)
        sid = source_id(slug, relative)
        source_ids[relative] = sid
        all_sources.append(
            {
                "source_id": sid,
                "path": relative,
                "title": title_from_path(path),
                "publisher": "",
                "published_at": date_from_filename(path),
                "language": "unknown",
                "source_type": source_type(path),
                "tlp": "TLP:CLEAR",
                "reliability": "medium",
                "sha256": None,
                "analyst_notes": "Repository source; publication metadata requires review.",
            }
        )
    known_source_paths = {source["path"] for source in all_sources}
    for relative in actor.get("reported_sources", []):
        if relative in known_source_paths:
            continue
        path = root / relative
        all_sources.append(
            {
                "source_id": source_id(slug, relative),
                "path": relative,
                "title": title_from_path(path),
                "publisher": "",
                "published_at": date_from_filename(path),
                "language": "unknown",
                "source_type": source_type(path),
                "tlp": "TLP:CLEAR",
                "reliability": "medium",
                "sha256": None,
                "analyst_notes": (
                    "Original report containing actor-associated evidence. "
                    "IOC ingestion uses the scoped evidence extract to avoid "
                    "cross-actor contamination."
                ),
            }
        )
    profile["sources"] = all_sources

    mitre_group = attack["groups"].get(actor.get("mitre_group_id", ""))
    if mitre_group:
        profile["actor"]["canonical_name"] = actor["name"]
        profile["actor"]["description"] = mitre_group.get("description", "")
        profile["actor"]["first_seen"] = time_point(
            mitre_group.get("first_seen"), "mitre-attack"
        )
        profile["actor"]["last_seen"] = time_point(
            mitre_group.get("last_seen"), "mitre-attack"
        )
    alias_names = []
    seen_alias_names: set[str] = set()
    for alias_name in [
        *actor.get("aliases", []),
        *(mitre_group.get("aliases", []) if mitre_group else []),
    ]:
        normalized_alias = normalized_name(alias_name)
        if (
            not normalized_alias
            or normalized_alias == normalized_name(actor["name"])
            or normalized_alias in seen_alias_names
        ):
            continue
        seen_alias_names.add(normalized_alias)
        alias_names.append(alias_name)
    profile["actor"]["aliases"] = [
        {
            "name": name,
            "vendor": "MITRE ATT&CK" if mitre_group and name in mitre_group.get("aliases", []) else "catalog",
            "scope": "overlapping",
            "confidence": "high" if mitre_group else "medium",
            "evidence_refs": [reference_source_id if mitre_group else workbook_source_id],
            "analyst_notes": "Alias scope must be reviewed before publication.",
        }
        for name in alias_names
        if normalized_name(name) != normalized_name(actor["name"])
    ]

    if workbook_record:
        mapped_names = actor_name_cells(workbook_record)
        existing = {normalized_name(item["name"]) for item in profile["actor"]["aliases"]}
        for name in mapped_names:
            if normalized_name(name) in existing | {normalized_name(actor["name"])}:
                continue
            profile["actor"]["aliases"].append(
                {
                    "name": name,
                    "vendor": "actor-mapping-workbook",
                    "scope": "unknown",
                    "confidence": "medium",
                    "evidence_refs": [workbook_source_id],
                    "analyst_notes": f"Workbook {workbook_record['sheet']} row {workbook_record['row']}; mapping requires review.",
                }
            )
            existing.add(normalized_name(name))
        if workbook_record["sheet"] in {
            "China", "Russia", "North Korea", "Iran", "Israel"
        }:
            country = workbook_record["sheet"]
            profile["attribution"].update(
                {
                    "countries": [country],
                    "sponsor_type": "state" if country != "Israel" or "private-sector-offensive" not in actor.get("actor_types", []) else "private-sector-offensive",
                    "assessment": f"The repository mapping workbook places this actor in the {country} worksheet.",
                    "confidence": "medium",
                    "evidence_refs": [workbook_source_id],
                    "analyst_notes": "Workbook attribution is a secondary mapping and must be corroborated.",
                }
            )

    motivation_types = []
    types = set(actor.get("actor_types", []))
    if "state-sponsored" in types:
        motivation_types.append(("espionage", "State-sponsored intelligence collection or strategic operations."))
    if "financially-motivated" in types or "business-email-compromise" in types:
        motivation_types.append(("financial-gain", "Financially motivated intrusion or fraud."))
    if "hacktivist-collective" in types:
        motivation_types.append(("ideological", "Ideological or political hacktivism."))
    if "private-sector-offensive" in types or "surveillance-vendor" in types:
        motivation_types.append(("commercial", "Commercial offensive-security or surveillance operations."))
    profile["motivations"] = [
        {
            "type": kind,
            "description": description,
            "confidence": "low",
            "evidence_refs": [workbook_source_id],
            "analyst_notes": "Inferred from catalog actor type; corroborate with actor-specific reporting.",
        }
        for kind, description in motivation_types
    ]

    software_items: list[dict[str, Any]] = []
    if mitre_group:
        for ref in mitre_group.get("software_refs", []):
            if ref in attack["software"]:
                software_items.append(attack["software"][ref])
    seen_software = set()
    for software in software_items:
        name = software["name"]
        seen_software.add(normalized_name(name))
        kind = software["software_type"]
        item = capability(
            "malware" if kind == "malware" else "tool",
            name,
            software.get("description", ""),
            [reference_source_id],
            aliases=software.get("aliases", []),
            types=software.get("platforms", []),
            confidence="high",
        )
        profile["capabilities"]["malware" if kind == "malware" else "tools"].append(item)
    for name in workbook_software(workbook_record):
        if normalized_name(name) in seen_software or len(name) > 120:
            continue
        profile["capabilities"]["malware"].append(
            capability(
                "malware",
                name,
                "The actor-mapping workbook lists this software or tool.",
                [workbook_source_id],
                confidence="medium",
            )
        )
        seen_software.add(normalized_name(name))

    operations = workbook_operations(workbook_record)
    seen_activity_ids: set[str] = set()
    for name in operations:
        activity_id = f"activity--{slugify(name)}"
        if activity_id in seen_activity_ids:
            continue
        seen_activity_ids.add(activity_id)
        profile["activities"].append(
            {
                "activity_id": activity_id,
                "name": name,
                "activity_type": "operation",
                "first_observed": unknown_time(),
                "last_observed": unknown_time(),
                "description": "Operation name listed in the repository actor-mapping workbook.",
                "target_refs": [],
                "malware_refs": [],
                "infrastructure_refs": [],
                "confidence": "medium",
                "evidence_refs": [workbook_source_id],
                "analyst_notes": "Operation-to-actor mapping requires source-level review.",
            }
        )
    if mitre_group:
        for ref in mitre_group.get("campaign_refs", []):
            campaign = attack["campaigns"].get(ref)
            if not campaign:
                continue
            activity_id = f"activity--{slugify(campaign['name'])}"
            if activity_id in seen_activity_ids:
                continue
            seen_activity_ids.add(activity_id)
            profile["activities"].append(
                {
                    "activity_id": activity_id,
                    "name": campaign["name"],
                    "activity_type": "campaign",
                    "first_observed": time_point(campaign.get("first_seen"), "mitre-attack"),
                    "last_observed": time_point(campaign.get("last_seen"), "mitre-attack"),
                    "description": campaign.get("description", ""),
                    "target_refs": [],
                    "malware_refs": [],
                    "infrastructure_refs": [],
                    "confidence": "high",
                    "evidence_refs": [reference_source_id],
                    "analyst_notes": "",
                }
            )

    target_text = workbook_target_text(workbook_record)
    countries, sectors = derive_targets(target_text, [workbook_source_id])
    profile["targets"]["countries"] = countries
    profile["targets"]["sectors"] = sectors
    profile["targets"]["analyst_notes"] = (
        "Structured targets are extracted from workbook prose and require review."
        if target_text
        else "No structured target statement was available in the mapping workbook."
    )

    technique_evidence: dict[str, dict[str, Any]] = {}
    if mitre_group:
        for technique_id in mitre_group.get("technique_ids", []):
            technique_evidence[technique_id] = {
                "evidence_refs": [reference_source_id],
                "context": "MITRE ATT&CK maps this technique to the actor.",
                "confidence": "high",
            }
    if scan_report_ttps:
        for technique_id, found in report_techniques(files, root, source_ids).items():
            record = technique_evidence.setdefault(
                technique_id,
                {"evidence_refs": [], "context": found["context"], "confidence": "medium"},
            )
            record["evidence_refs"] = sorted(
                set(record["evidence_refs"]) | set(found["evidence_refs"])
            )
            if not record.get("context"):
                record["context"] = found["context"]

    for technique_id, evidence in sorted(technique_evidence.items()):
        technique = attack["techniques"].get(technique_id, {})
        tactics = technique.get("tactics", []) or ["Uncategorized"]
        profile["ttps"].append(
            {
                "ttp_id": f"ttp--{technique_id.lower().replace('.', '-')}--general",
                "tactic": ", ".join(tactics),
                "technique_id": technique_id,
                "technique_name": technique.get("name", f"MITRE ATT&CK {technique_id}"),
                "observed_behavior": evidence.get("context", ""),
                "activity_refs": [],
                "malware_refs": [],
                "infrastructure_refs": [],
                "first_observed": unknown_time(),
                "last_observed": unknown_time(),
                "confidence": evidence["confidence"],
                "evidence_refs": evidence["evidence_refs"],
                "analyst_notes": "General actor-level mapping; split by activity when source evidence permits.",
            }
        )

    capability_names = [
        item["name"]
        for kind in ("malware", "tools")
        for item in profile["capabilities"][kind]
    ]
    profile["diamond_model"].update(
        {
            "adversary": profile["actor"]["description"],
            "capability": ", ".join(capability_names),
            "infrastructure": "",
            "victim": target_text,
            "socio_political": ", ".join(profile["attribution"]["countries"]),
            "analyst_notes": "Bootstrap model; infrastructure and campaign-level pivots require report review.",
        }
    )
    profile["free_text"].update(
        {
            "executive_summary": (
                f"{actor['name']}の標準化プロファイル。"
                f"リポジトリ内の専用資料{len(files)}件とMITRE ATT&CK、"
                "アクターマッピング表を基礎情報としている。"
            ),
            "history": "; ".join(operations),
            "capability_details": ", ".join(capability_names),
            "infrastructure_details": "",
            "targeting_details": target_text,
            "additional_notes": "自動構造化した項目はdraftであり、candidateとunknownを分析者がレビューする。",
        }
    )
    profile["assessment"].update(
        {
            "gaps": [
                "Unknown observation dates must not be replaced by publication dates.",
                "Automatically mapped aliases, targets, and workbook software require analyst review.",
            ],
            "uncertainties": [
                "Vendor cluster boundaries may differ from the canonical name used here."
            ],
            "collection_notes": f"Catalog source roots: {', '.join(actor['source_dirs'])}",
        }
    )

    activity_names = {
        item["activity_id"]: normalized_name(item["name"])
        for item in profile["activities"]
    }
    malware_names = {
        item["id"]: normalized_name(item["name"])
        for item in profile["capabilities"]["malware"]
    }
    manifest_sources = []
    for source in all_sources[2:]:
        if source["path"] not in ingest_source_paths:
            continue
        path_norm = normalized_name(source["path"])
        manifest_sources.append(
            {
                "source_id": source["source_id"],
                "path": source["path"],
                "published_at": source["published_at"],
                "campaign_refs": [
                    ref for ref, name in activity_names.items() if len(name) >= 5 and name in path_norm
                ],
                "malware_refs": [
                    ref for ref, name in malware_names.items() if len(name) >= 5 and name in path_norm
                ],
                "infrastructure_refs": [],
                "roles": [],
                "confidence": source["reliability"],
                "allow_plain_domains": False,
            }
        )
    manifest = {
        "schema_version": "1.0.0",
        "actor_ref": profile["profile_id"],
        "repository_root": "../..",
        "defaults": {
            "confidence": "medium",
            "tlp": "TLP:CLEAR",
            "campaign_refs": [],
            "malware_refs": [],
            "infrastructure_refs": [],
            "roles": [],
            "default_observed_at": unknown_time(),
        },
        "sources": manifest_sources,
        "source_groups": [],
    }
    return profile, manifest


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--catalog",
        type=Path,
        default=FRAMEWORK_DIR / "corpus-catalog.json",
    )
    parser.add_argument("--repository-root", type=Path, default=Path.cwd())
    parser.add_argument("--profiles-root", type=Path, default=Path("profiles"))
    parser.add_argument("--actor", action="append", help="only process this slug; repeatable")
    parser.add_argument("--exclude-actor", action="append", default=[])
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--scan-report-ttps", action="store_true")
    args = parser.parse_args()

    root = args.repository_root.resolve()
    catalog = load_json(args.catalog.resolve())
    attack = load_json((root / catalog["reference_sources"]["mitre_attack_index"]).resolve())
    workbook_rows = load_workbook_rows(
        (root / catalog["reference_sources"]["actor_mapping_workbook"]).resolve()
    )
    wanted = set(args.actor or [])
    excluded = set(args.exclude_actor)
    results = []
    for actor in catalog["actors"]:
        if wanted and actor["slug"] not in wanted:
            continue
        if actor["slug"] in excluded:
            continue
        output_dir = (args.profiles_root.resolve() / actor["slug"])
        profile_path = output_dir / "actor-profile.json"
        manifest_path = output_dir / "ioc-sources.json"
        if profile_path.exists() and not args.overwrite:
            results.append({"actor": actor["slug"], "status": "preserved-existing"})
            continue
        files = expand_actor_files(root, actor["source_dirs"])
        record = find_workbook_record(actor, workbook_rows)
        profile, manifest = create_profile(
            actor,
            root,
            files,
            attack,
            record,
            scan_report_ttps=args.scan_report_ttps,
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "generated").mkdir(exist_ok=True)
        write_json_atomic(profile_path, profile)
        write_json_atomic(manifest_path, manifest)
        results.append(
            {
                "actor": actor["slug"],
                "status": "created",
                "sources": len(files),
                "aliases": len(profile["actor"]["aliases"]),
                "software": len(profile["capabilities"]["malware"])
                + len(profile["capabilities"]["tools"]),
                "ttps": len(profile["ttps"]),
                "workbook_match": bool(record),
            }
        )
        print(json.dumps(results[-1], ensure_ascii=False))
    summary = {
        "actors_selected": len(results),
        "created": sum(item["status"] == "created" for item in results),
        "preserved_existing": sum(item["status"] == "preserved-existing" for item in results),
    }
    print(json.dumps({"summary": summary}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
